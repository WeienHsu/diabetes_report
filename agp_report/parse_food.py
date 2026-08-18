"""解析 Google Sheet「飲食日誌」的匯出檔，CSV 或 xlsx 皆可。

鄰近的多筆明細視為同一次進食（例如晚餐的主菜與白飯分兩列、或邊吃邊補記
而差了幾分鐘），營養素加總、品項合併成一份餐次。不合併的話同一頓飯會
拆成數張餐後反應卡，共用同一劑胰島素、重複計算同一個血糖峰值。

手機上的 Google Sheets 匯不出單頁 CSV，只能存整本 xlsx。整本反而更安全——
CSV 只會拿到當前分頁（停在「食物資料庫」就解析出 0 餐），xlsx 兩頁都在，
可以按名字挑對的那一頁。
"""

import csv
from dataclasses import dataclass, field
from datetime import datetime, timedelta

# Sheet 的時間欄位小時不補零（例如 "2026-08-14 8:24"）；xlsx 存的若是真正的
# 日期值，openpyxl 會直接給 datetime，只有存成文字時才走這兩個格式。
TIMESTAMP_FMT = "%Y-%m-%d %H:%M"
TIMESTAMP_FMT_SEC = "%Y-%m-%d %H:%M:%S"

# xlsx 活頁簿裡要找的分頁；找不到就退而求其次挑有「日期時間」欄的那一頁
# （使用者改過分頁名稱時仍能用，而「食物資料庫」沒有這一欄，不會選錯）。
SHEET_NAME = "飲食日誌"
TIME_COLUMN = "日期時間"

# 距離同一餐起始多久內的紀錄併入該餐
CLUSTER_WINDOW = timedelta(minutes=30)


@dataclass
class Item:
    label: str          # 該列原本的餐別，合併後仍要能分行顯示
    brand: str
    name: str
    serving: str
    carbs: float
    protein: float
    fat: float
    kcal: float


@dataclass
class Meal:
    when: datetime                   # 第一筆進食紀錄
    last: datetime | None = None     # 最後一筆；合併餐次才會與 when 不同
    items: list[Item] = field(default_factory=list)

    @property
    def ends(self) -> datetime:
        return self.last or self.when

    @property
    def carbs(self) -> float:
        return sum(i.carbs for i in self.items)

    @property
    def protein(self) -> float:
        return sum(i.protein for i in self.items)

    @property
    def fat(self) -> float:
        return sum(i.fat for i in self.items)

    @property
    def kcal(self) -> float:
        return sum(i.kcal for i in self.items)

    @property
    def groups(self) -> list[tuple[str, list[Item]]]:
        """依原始餐別分組，順序照時間。

        20:05 的晚餐與 20:30 的點心會合併成同一次進食（血糖反應分不開），
        但卡片仍要分行標出哪些品項屬於哪一餐別——否則整張卡標著「晚餐」，
        裡面卻有一半碳水來自使用者明確記為「點心」的東西。
        """
        out: list[tuple[str, list[Item]]] = []
        for item in self.items:
            label = item.label or "未指定"
            if out and out[-1][0] == label:
                out[-1][1].append(item)
            else:
                out.append((label, [item]))
        return out

    @property
    def label(self) -> str:
        """精簡場合，例如漲幅排行與每日詳圖的數值框。"""
        names = []
        for label, _items in self.groups:
            if label and label != "未指定" and label not in names:
                names.append(label)
        return "＋".join(names) or "未指定"

    @property
    def title(self) -> str:
        return "＋".join(i.name for i in self.items)


class FoodFormatError(Exception):
    """檔案本身無法當成飲食日誌讀。訊息寫給使用者看。"""


def _num(raw) -> float:
    """xlsx 的數值欄位拿到的是 int/float，CSV 一律是字串。"""
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(raw.strip())
    except (ValueError, AttributeError):
        return 0.0


def _text(raw) -> str:
    return "" if raw is None else str(raw).strip()


def _when(raw) -> datetime | None:
    if isinstance(raw, datetime):
        return raw
    for fmt in (TIMESTAMP_FMT, TIMESTAMP_FMT_SEC):
        try:
            return datetime.strptime(_text(raw), fmt)
        except ValueError:
            continue
    return None


def is_xlsx(path: str) -> bool:
    """認檔頭而不是副檔名——手機上傳的檔名不保證帶得對。"""
    with open(path, "rb") as fh:
        return fh.read(2) == b"PK"


def _rows_csv(path: str):
    with open(path, encoding="utf-8-sig", newline="") as fh:
        yield from csv.DictReader(fh)


def _rows_xlsx(path: str):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:      # BadZipFile、InvalidFileException 等
        raise FoodFormatError("這個檔案不是可讀的 Excel 活頁簿，也不是 CSV。"
                              "請從 Google Sheets 重新匯出。") from exc
    try:
        sheets = [wb[SHEET_NAME]] if SHEET_NAME in wb.sheetnames else list(wb.worksheets)
        for ws in sheets:
            rows = ws.iter_rows(values_only=True)
            header = [_text(c) for c in next(rows, ())]
            if TIME_COLUMN not in header:
                continue
            for row in rows:
                yield dict(zip(header, row))
            return
        raise FoodFormatError(
            f"這本活頁簿裡找不到有「{TIME_COLUMN}」欄的分頁"
            f"（分頁：{'、'.join(wb.sheetnames)}）。"
            "請確認匯出的是「飲食與三大營養素紀錄表」。")
    finally:
        wb.close()


def parse(path: str) -> list[Meal]:
    entries: list[tuple[datetime, Item]] = []
    for row in (_rows_xlsx(path) if is_xlsx(path) else _rows_csv(path)):
        when = _when(row.get(TIME_COLUMN))
        if when is None:
            continue
        entries.append((when, Item(
            label=_text(row.get("餐別")),
            brand=_text(row.get("品牌")),
            name=_text(row.get("產品名稱")),
            serving=_text(row.get("攝取份量")),
            carbs=_num(row.get("碳水化合物 (g)")),
            protein=_num(row.get("蛋白質 (g)")),
            fat=_num(row.get("脂肪 (g)")),
            kcal=_num(row.get("熱量 (kcal)")),
        )))

    meals: list[Meal] = []
    for when, item in sorted(entries, key=lambda e: e[0]):
        if meals and when - meals[-1].when <= CLUSTER_WINDOW:
            meals[-1].items.append(item)
            meals[-1].last = when
        else:
            meals.append(Meal(when=when, last=when, items=[item]))
    return meals
